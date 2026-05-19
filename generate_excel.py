import os
import json
import re
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(BASE_DIR, "scraped_results", "scraped_data.json")
OUTPUT_EXCEL = os.path.join(BASE_DIR, "scraped_results", "sikkim_policy_files_index.xlsx")
PDF_DIR = os.path.join(BASE_DIR, "sikkim_policy_briefs")

def get_file_size_display(size_in_bytes):
    if size_in_bytes < 1024:
        return f"{size_in_bytes} B"
    elif size_in_bytes < 1024 * 1024:
        return f"{size_in_bytes / 1024:.2f} KB"
    else:
        return f"{size_in_bytes / (1024 * 1024):.2f} MB"

def clean_title(title):
    if not title:
        return "Untitled Document"
    # Remove leading/trailing brackets or spaces (like [PDF] prefix)
    title = re.sub(r"^\[PDF\]\s*", "", title)
    return title.strip()

def extract_year(pub_info):
    if not pub_info or not isinstance(pub_info, str):
        return None
    # Look for a 4-digit number between 1800 and 2099
    match = re.search(r"\b(18\d{2}|19\d{2}|20\d{2})\b", pub_info)
    if match:
        return match.group(1)
    return None

def main():
    print("=========================================================")
    print("Generating Professional Excel File from Scraped Data")
    print("=========================================================")

    # ---------------------------------------------------------
    # PART 1: Parse scraped_data.json (Remote/Scraped Files)
    # ---------------------------------------------------------
    if not os.path.exists(JSON_PATH):
        print(f"Error: Scraped data file not found at {JSON_PATH}")
        return

    with open(JSON_PATH, "r", encoding="utf-8") as f:
        scraped_data = json.load(f)

    scraped_rows = []
    for idx, item in enumerate(scraped_data, 1):
        url = item.get("url") or ""
        title = item.get("title") or ""
        source = item.get("source") or ""
        keyword = item.get("keyword") or ""
        pub_info = item.get("publication_info") or ""
        scraped_at = item.get("scraped_at") or ""

        # Determine Name
        name = clean_title(title)

        # Determine Item Type and Extension
        is_pdf = "pdf" in url.lower() or "pdf" in title.lower()
        if is_pdf:
            item_type = "PDF Document"
            ext = ".pdf"
        elif source == "Google Scholar":
            item_type = "Academic Article"
            ext = ".html"
        else:
            item_type = "Government Web Page"
            ext = ".html"

        # Check if URL has a specific extension
        url_clean = url.split("?")[0].split("#")[0]
        match_ext = re.search(r"\.([a-zA-Z0-9]+)$", url_clean)
        if match_ext and not is_pdf:
            ext = "." + match_ext.group(1)
            # Limit extension length to prevent long junk strings
            if len(ext) > 5:
                ext = ".html"

        # Determine Modified Date
        year = extract_year(pub_info)
        if year:
            modified = year
        elif scraped_at:
            try:
                dt = datetime.fromisoformat(scraped_at)
                modified = dt.strftime("%Y-%m-%d")
            except Exception:
                modified = scraped_at[:10]
        else:
            modified = "N/A"

        # Parent Folder (we group them by the keyword/topic they belong to)
        parent_folder = keyword.replace("sikkim ", "").title()

        scraped_rows.append({
            "S.No": idx,
            "Relative Path": url,
            "Name": name,
            "Item Type": item_type,
            "Extension": ext,
            "Size": "Remote Link",
            "Modified": modified,
            "Parent Folder": parent_folder
        })

    df_scraped = pd.DataFrame(scraped_rows)

    # ---------------------------------------------------------
    # PART 2: Parse Local Files (Workspace PDFs and scripts)
    # ---------------------------------------------------------
    local_rows = []
    local_idx = 1

    # Add PDFs from local directories (both sikkim_policy_briefs and scraped_policy_pdfs)
    local_dirs = [PDF_DIR, os.path.join(BASE_DIR, "scraped_policy_pdfs")]
    for l_dir in local_dirs:
        if os.path.exists(l_dir):
            for root, dirs, files in os.walk(l_dir):
                # Sort files to keep output ordered
                for file in sorted(files):
                    file_path = os.path.join(root, file)
                    rel_path = os.path.relpath(file_path, BASE_DIR)
                    name, ext = os.path.splitext(file)
                    size_bytes = os.path.getsize(file_path)
                    mtime = os.path.getmtime(file_path)
                    modified_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
                    parent = os.path.basename(root)

                    local_rows.append({
                        "S.No": local_idx,
                        "Relative Path": rel_path.replace("\\", "/"),
                        "Name": name,
                        "Item Type": "PDF Document",
                        "Extension": ext,
                        "Size": get_file_size_display(size_bytes),
                        "Modified": modified_str,
                        "Parent Folder": parent
                    })
                    local_idx += 1


    # Add other core project files for completeness
    project_files = ["scraping.py", "requirements.txt", ".gitignore", "README.md"]
    for file in project_files:
        file_path = os.path.join(BASE_DIR, file)
        if os.path.exists(file_path):
            name, ext = os.path.splitext(file)
            size_bytes = os.path.getsize(file_path)
            mtime = os.path.getmtime(file_path)
            modified_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")

            # Determine item type
            if ext == ".py":
                item_type = "Python Source Code"
            elif ext == ".txt":
                item_type = "Text Document"
            elif ext == ".md":
                item_type = "Markdown Document"
            else:
                item_type = "Configuration File"

            local_rows.append({
                "S.No": local_idx,
                "Relative Path": file,
                "Name": file if not name else name,
                "Item Type": item_type,
                "Extension": ext,
                "Size": get_file_size_display(size_bytes),
                "Modified": modified_str,
                "Parent Folder": "Project Root"
            })
            local_idx += 1

    df_local = pd.DataFrame(local_rows)

    # ---------------------------------------------------------
    # PART 3: Write to Excel and Style with OpenPyXL
    # ---------------------------------------------------------
    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        df_scraped.to_excel(writer, sheet_name="Scraped Files Index", index=False)
        df_local.to_excel(writer, sheet_name="Local Files Index", index=False)

    # Re-open the file with openpyxl to apply premium styles
    wb = openpyxl.load_workbook(OUTPUT_EXCEL)

    # Styling Elements
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Teal/Blue
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="000000")
    link_font = Font(name="Segoe UI", size=10, color="0563C1", underline="single")
    sno_font = Font(name="Segoe UI", size=10, bold=True, color="555555")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    zebra_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid") # Very light blue tint
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True

        # Header height and styling
        ws.row_dimensions[1].height = 28

        # Format Headers
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Format Data Rows
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            is_even = (row_idx % 2 == 0)
            current_fill = zebra_fill if is_even else white_fill

            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = current_fill
                cell.border = thin_border
                
                # Check column headers
                header_name = ws.cell(row=1, column=col_idx).value

                # Default alignments & fonts
                if header_name == "S.No":
                    cell.font = sno_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif header_name == "Extension":
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif header_name in ["Size", "Modified"]:
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif header_name == "Relative Path":
                    val = cell.value
                    if val and (val.startswith("http://") or val.startswith("https://")):
                        cell.font = link_font
                        cell.hyperlink = val
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.font = data_font
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-fit Column Widths with padding
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            header_val = col[0].value

            for cell in col:
                val_str = str(cell.value or '')
                # If it's a URL, don't let it expand the column too much, cap it at 40 chars
                if header_val == "Relative Path" and (val_str.startswith("http://") or val_str.startswith("https://")):
                    val_len = min(len(val_str), 40)
                else:
                    val_len = len(val_str)
                if val_len > max_len:
                    max_len = val_len
            
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(OUTPUT_EXCEL)
    print(f"Success! Excel sheet created and styled at: {OUTPUT_EXCEL}")

if __name__ == "__main__":
    main()
